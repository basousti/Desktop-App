import sys 
from PyQt5 import QtWidgets , QtCore
from PyQt5.QtGui import QIcon , QPixmap ,QPalette,QFont,QColor
import openpyxl
import os
import pandas as pd 

class Personne:
    def __init__(self, nom,prenom,CIN, age,adresse,telephon, ville, date_infection,décé):
        self.nom = nom
        self.prenom = prenom
        self.CIN = CIN
        self.age = age
        self.adresse = adresse
        self.telephon=telephon
        self.ville = ville
        self.date_infection = date_infection
        self.décé = décé

class maladies:
    def __init__(self ,CIN,code,nomal,nb):
        self.nomal=nomal
        self.CIN=CIN
        self.code=code
        self.nb = nb

class my_window(QtWidgets.QMainWindow): #Instead of using functions, we can easily design desktop applications and avoid errors by using classes.
    def __init__(self):
        super(my_window, self).__init__()

        self.personne =[]
        self.Maladie =[]

        self.setGeometry(1200, 50, 700, 700)
        self.setWindowTitle("CORONA virus")
        self.setWindowIcon(QIcon("ISIMM.jpg"))
        self.setToolTip("clicker")#when you don't move the mouse the text appears
        
        background =QtWidgets.QLabel(self)
        pixmap =QPixmap("coro.jpg").scaled(1990, 980)
        background.setPixmap(pixmap)
        #background.setGeometry(0, 0, self.width(), self.height())
        background.setGeometry(0, 0, 1000000, 990)
        font = QFont("Century Gothic", 10)
        palette = QPalette()
        palette.setColor(QPalette.WindowText, QColor(255, 0, 0)) # Rouge
        font.setBold(True)
        self.setFont(font)
        self.setPalette(palette)
        shadow = QFont()
        shadow.setHintingPreference(QFont.PreferNoHinting)


        menubar = self.menuBar()
        menu_bar = menubar.addMenu('Menu personne')
        
        menu_bar.addSeparator()
        add1=menu_bar.addMenu('Mise a jours des personnes')
        AE=QtWidgets.QAction('Ajouter personne',self)
        AE.triggered.connect(self.ident)
        add1.addAction(AE)

        add1.addSeparator()
        SE=add1.addMenu('Supprimer personne')
        S1=QtWidgets.QAction('suppression personne donné',self)
        S1.triggered.connect(self.suppression_personne)
        S2=QtWidgets.QAction('suppression des personnes d\'une nationaliter ',self)
        S2.triggered.connect(self.Supprimer_N)
        S3=QtWidgets.QAction('suppression des personnes d\'un indicatif',self)
        S3.triggered.connect(self.Supprimer_I)
        SE.addAction(S1)
        SE.addAction(S2)
        SE.addAction(S3)

        add1.addSeparator()
        M1=QtWidgets.QAction('Modifier personne',self)
        M1.triggered.connect(self.Modifier)
        add1.addAction(M1)

        menu_bar.addSeparator()
        add2=menu_bar.addMenu('Recherche et Affichage')
        one=QtWidgets.QAction('Contenue de dictionnaire personne',self)
        one.triggered.connect(self.afficher_contenu_personnes)
        two=QtWidgets.QAction('Recherche par numéro télèphone',self)
        two.triggered.connect(self.rechercher_personne_tele)
        three=QtWidgets.QAction('Recherch par indicatif',self)
        three.triggered.connect(self.Rech_IN) 
        four=QtWidgets.QAction('Recherche par nationalité ',self)
        four.triggered.connect(self.rechercher_personne_nationalite)
        five=QtWidgets.QAction('Recherche des personnes décédés',self)
        five.triggered.connect(self.rechercher_personne_decede)
        six=QtWidgets.QAction('Recherche des personnes non décédés',self)
        six.triggered.connect(self.rechercher_personne_non_decede)
        add2.addAction(one)
        add2.addAction(two)
        add2.addAction(three)
        add2.addAction(four)
        add2.addAction(five)
        add2.addAction(six)

        add1.addSeparator()
        exit_action = QtWidgets.QAction('EXIT', self)
        exit_action.triggered.connect(self.close)
        menu_bar.addAction(exit_action)

        menu_maladies = menubar.addMenu('Gestion des maladies')
        menu_maladies.addSeparator()
        add3 = menu_maladies.addMenu('Mise à jour')
        A = QtWidgets.QAction('Ajouter une nouvelle maladie', self)
        A.triggered.connect(self.mal)
        add3.addAction(A)
        S= QtWidgets.QAction('supprimer une maladie', self)
        S.triggered.connect(self.suppression_maladie)
        add3.addAction(S)
        MEE = QtWidgets.QAction("Modifier les données d'une maladie",self)
        MEE.triggered.connect(self.ANN_DC)
        add3.addAction(MEE)
       
        add4 = menu_maladies.addMenu('Recherche et Affichage')
        one = QtWidgets.QAction('Contenue de dictionnaire maladies', self)
        one.triggered.connect(self.afficher_contenu_maladie)
        two = QtWidgets.QAction('Recherche par une maladie', self)
        two.triggered.connect(self.rechercher_par_maladie)
        three = QtWidgets.QAction("Recherche maladie d'une personne ", self)
        three.triggered.connect(self.rechercher_maladie_personne)
        four = QtWidgets.QAction('Recherche le pourcentage de chaque maladie', self)
        four.triggered.connect(self.rechercher_pourcentage_maladies)
        five = QtWidgets.QAction('Recherche maladie de chaque personne', self)
        five.triggered.connect(self.afficher_maladies_chaque_personne)
        add4.addAction(one)
        add4.addAction(two)
        add4.addAction(three)
        add4.addAction(four)
        add4.addAction(five)

        exit_action = QtWidgets.QAction('EXIT', self)
        exit_action.triggered.connect(self.close)
        menu_maladies.addAction(exit_action)
        
        menu_calcul = menubar.addMenu('Calcul et affichage')
        a=QtWidgets.QAction('afficher par nationalité',self)
        a.triggered.connect(self.afficher_personnes_par_ville)
        b=QtWidgets.QAction('Personnes en quarantine',self)
        c=QtWidgets.QAction('Personnes décés',self)
        d=QtWidgets.QAction('Personnes à risque',self)
        menu_calcul.addAction(a)
        menu_calcul.addAction(b)
        menu_calcul.addAction(c)
        menu_calcul.addAction(d)
        menu_enregistrement = menubar.addMenu('Enregistrement Et Récupération ')
        x=QtWidgets.QAction('enregistrement fichier Personnes',self)
        x.triggered.connect(self.enregistrer_personne_excel)
        m=QtWidgets.QAction('Récupération fichier Personnes',self)
        m.triggered.connect(self.recuperer_personne_excel)
        w=QtWidgets.QAction('enregistrement fichier Maladies',self)
        w.triggered.connect(self.enregistrer_maladie_excel)
        z=QtWidgets.QAction('recuperation fichier Maladies',self)
        z.triggered.connect(self.recuperer_maladie_excel)
        menu_enregistrement.addAction(x)
        menu_enregistrement.addAction(m)
        menu_enregistrement.addAction(w)
        menu_enregistrement.addAction(z)

    def ident(self):
        
        self.centralWidget =QtWidgets.QWidget()
        self.setCentralWidget(self.centralWidget)

        self.layout =QtWidgets.QVBoxLayout()

        self.label_nom =QtWidgets.QLabel("Nom :")
        self.edit_nom =QtWidgets.QLineEdit()
        self.edit_nom.setMaximumWidth(350)

        self.label_prenom = QtWidgets.QLabel("Prénom :")
        self.edit_prenom = QtWidgets.QLineEdit()
        self.edit_prenom.setMaximumWidth(350)

        self.label_CIN = QtWidgets.QLabel("Num CIN : ")
        self.edit_CIN = QtWidgets.QLineEdit()
        self.edit_CIN.setMaximumWidth(350)

        self.label_age = QtWidgets.QLabel("Âge :")
        self.edit_age = QtWidgets.QLineEdit()
        self.edit_age.setMaximumWidth(350)

        self.label_adrs = QtWidgets.QLabel("Adress :")
        self.edit_adrs = QtWidgets.QLineEdit()
        self.edit_adrs.setMaximumWidth(350)

        self.label_tele = QtWidgets.QLabel("Téléphone :")
        self.edit_tele = QtWidgets.QLineEdit()
        self.edit_tele.setMaximumWidth(350)

        self.label_ville = QtWidgets.QLabel("Nationalité :")
        self.edit_ville = QtWidgets.QLineEdit()
        self.edit_ville.setMaximumWidth(350)

        self.label_date_infection = QtWidgets.QLabel("Date d'infection :")
        self.edit_date_infection = QtWidgets.QDateEdit()
        self.edit_date_infection.setMinimumDate(QtCore.QDate.currentDate().addYears(-100))
        self.edit_date_infection.setMaximumDate(QtCore.QDate.currentDate())
        self.submit_button = QtWidgets.QPushButton('Soumettre', self)
        self.edit_date_infection.setMaximumWidth(350)

        self.décé = QtWidgets.QLabel(" Cette personne est elle Décédée ? : ")
        self.res1= QtWidgets.QRadioButton("True")
        self.res2= QtWidgets.QRadioButton("False")
       
        self.btn_ajouter = QtWidgets.QPushButton("Ajouter",self)
        self.btn_ajouter.clicked.connect(self.ajouter_personne)
        self.btn_ajouter.setMaximumWidth(350)
        self.btn_ajouter.setStyleSheet("background-color: Crimson; color: Gold; font-weight: Bold;")
    

        self.table_personnes = QtWidgets.QTableWidget()
        self.table_personnes.setColumnCount(9)
        self.table_personnes.setHorizontalHeaderLabels(["Nom", "Prénom", "CIN","Âge", "Adresse","Téléphone", "Nationalité", "Date d'infection","décé"])
        
        #layout permet de positionner dans l'interface 
        self.layout.addWidget(self.label_nom)
        self.layout.addWidget(self.edit_nom)
        self.layout.addWidget(self.label_prenom)
        self.layout.addWidget(self.edit_prenom)
        self.layout.addWidget(self.label_CIN)
        self.layout.addWidget(self.edit_CIN)
        self.layout.addWidget(self.label_age)
        self.layout.addWidget(self.edit_age)
        self.layout.addWidget(self.label_adrs)
        self.layout.addWidget(self.edit_adrs)
        self.layout.addWidget(self.label_tele)
        self.layout.addWidget(self.edit_tele)
        self.layout.addWidget(self.label_ville)
        self.layout.addWidget(self.edit_ville)
        self.layout.addWidget(self.label_date_infection)
        self.layout.addWidget(self.edit_date_infection)
        self.layout.addWidget(self.décé)
        self.layout.addWidget(self.res1)
        self.layout.addWidget(self.res2)
        self.layout.addWidget(self.btn_ajouter)
        self.layout.addWidget(self.table_personnes)

        self.centralWidget.setLayout(self.layout)
        self.table_personnes.setStyleSheet("background-color: Gray; color: FireBrick; font-weight: Bold;")        
        self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)


    def ajouter_personne(self):
        nom = self.edit_nom.text()
        prenom = self.edit_prenom.text()
        CIN = self.edit_CIN.text()
        age = self.edit_age.text()
        adresse = self.edit_adrs.text()
        telephon=self.edit_tele.text()
        ville = self.edit_ville.text()
        date_infection = self.edit_date_infection.text()
        if self.res1.isChecked():
            décé=self.res1.text()
        else:
            décé=self.res2.text()


        if nom and prenom and CIN and age and adresse and ville and telephon and date_infection and décé :
            personne = Personne(nom, prenom,CIN, age,adresse, telephon,ville,date_infection,décé)
            self.personne.append(personne)
            self.afficher_personnes()
            self.edit_nom.setText("")
        
            if nom.isalpha() ==False:
                QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le nom soit une chaine")
                sys.exit()
            self.edit_prenom.setText("")
            if prenom.isalpha() == False:
                QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le prénom soit une chaine")
                sys.exit()
            cin1=self.edit_CIN.setText("")
            if  CIN.isnumeric() == False:
                QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le CIN soit un entier")
                sys.exit()
            self.edit_age.setText("")
            if age.isnumeric() == False:
                QtWidgets.QMessageBox.warning(self, "Attention", "il faut que l'age soit un entier")
                sys.exit()
            self.edit_adrs.setText("")
            self.edit_tele.setText("")
            if telephon.isnumeric() == False:
                QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le num de téléphone soit un entier")
                sys.exit()
            self.edit_ville.setText("")
            self.edit_date_infection.setDate(QtCore.QDate.currentDate())
            self.décé.setText("")  
            
        else:
            QtWidgets.QMessageBox.warning(self, "Attention", "Veuillez remplir tous les champs.")

        wb = openpyxl.load_workbook("CORONA.xlsx")
        sheet = wb['F1']
        sheet["A1"] = "Name"
        sheet["B1"] = "Prénom"
        sheet["C1"]= "CIN"
        sheet["D1"]="Age"
        sheet["E1"]="Adresse"
        sheet["F1"]="Téléphone"
        sheet["G1"]="Nationalité"
        sheet["H1"]="Date d'infection"
        sheet["I1"]="Décé"
        
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=nom)
        sheet.cell(row=row, column=2, value=prenom)
        sheet.cell(row=row, column=3, value=CIN)
        sheet.cell(row=row, column=4, value=age)
        sheet.cell(row=row, column=5, value=adresse)
        sheet.cell(row=row, column=6, value=telephon)
        sheet.cell(row=row, column=7, value=ville)
        sheet.cell(row=row, column=8, value=date_infection)
        sheet.cell(row=row, column=9, value=décé)

        wb.save("CORONA.xlsx")
        
    def afficher_personnes(self):
       self.table_personnes.setRowCount(len(self.personne))
       for i, personne in enumerate(self.personne):
        nom = QtWidgets.QTableWidgetItem(personne.nom)
        prenom = QtWidgets.QTableWidgetItem(personne.prenom)
        CIN = QtWidgets.QTableWidgetItem(personne.CIN)
        age = QtWidgets.QTableWidgetItem(str(personne.age)) # Convert age to a string before setting it as QTableWidgetItem
        adresse = QtWidgets.QTableWidgetItem(personne.adresse)
        telephon = QtWidgets.QTableWidgetItem(personne.telephon)
        ville = QtWidgets.QTableWidgetItem(personne.ville)
        date_infection = QtWidgets.QTableWidgetItem(personne.date_infection)
        Deces = QtWidgets.QTableWidgetItem(str(personne.décé)) # Convert deces to a string before setting it as QTableWidgetItem
        
         # Set the QTableWidgetItem objects in the table (affichage dans le tableau)

        self.table_personnes.setItem(i, 0, nom)
        self.table_personnes.setItem(i, 1, prenom)
        self.table_personnes.setItem(i, 2, CIN)
        self.table_personnes.setItem(i, 3, age)
        self.table_personnes.setItem(i, 4, adresse)
        self.table_personnes.setItem(i, 5, telephon)
        self.table_personnes.setItem(i, 6, ville)
        self.table_personnes.setItem(i, 7, date_infection)
        self.table_personnes.setItem(i, 8, Deces)

    def suppression_personne(self):
        self.nnom = QtWidgets.QLabel("Nom :")
        self.Nom = QtWidgets.QLineEdit()
        self.prenom = QtWidgets.QLabel("Prénom :")
        self.Prenom = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.nnom)
        layout.addWidget(self.Nom)
        layout.addWidget(self.prenom)
        layout.addWidget(self.Prenom)
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Supprimer personne")
        dialog.setStyleSheet("background-color: #E1D0C0;color: Maroon; font-weight: Bold;")
        dialog.setLayout(layout)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)

        layout.addWidget(buttonBox)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            nom = self.Nom.text()
            prenom = self.Prenom.text()
            for i, personne in enumerate(self.personne):
                if personne.nom == nom and personne.prenom == prenom :
                    self.personne.pop(i)
                    self.table_personnes.removeRow(i)
                    break
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            for row in range(2, sheet.max_row +1):
                cell_nom = sheet.cell(row,1)
                cell_prenom = sheet.cell(row,2)
                if cell_nom.value  == nom and cell_prenom.value == prenom:
                    sheet.delete_rows(row)
                    wb.save("CORONA.xlsx")
                    QtWidgets.QMessageBox.information(self, "Succès", "Succèe de supression")
                    return   
        QtWidgets.QMessageBox.critical(self, "Error", "Aucune personne n'a été trouvée ")
        

    def Supprimer_N(self):
        self.ville = QtWidgets.QLabel("Nationalité :")
        self.Ville1 = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.ville)
        layout.addWidget(self.Ville1)
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Supprimer personne :")
        dialog.setLayout(layout)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        dialog.setStyleSheet("background-color: #E1D0C0;color: Maroon; font-weight: Bold;")
        layout.addWidget(buttonBox)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            NAS = self.Ville1.text()
            for i, personne in enumerate(self.personne):
                if personne.ville == NAS :
                    self.personne.pop(i)
                    self.table_personnes.removeRow(i)
                    break
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            for row in range(2, sheet.max_row +1):
                cell_NAS = sheet.cell(row,7)
                if cell_NAS.value  == NAS :
                    sheet.delete_rows(row)
                    wb.save("CORONA.xlsx")
                    QtWidgets.QMessageBox.information(self, "Succès", "Succès de suppression")
                    return
        QtWidgets.QMessageBox.critical(self, "Error", "Aucune personne n'a été trouvée avec cette nationalité")

    def Supprimer_I (self):
        self.tele = QtWidgets.QLabel("téléphone :")
        self.Tele = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.tele)
        layout.addWidget(self.Tele)
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Supprimer personne :")
        dialog.setLayout(layout)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        dialog.setStyleSheet("background-color: #E1D0C0;color: Maroon; font-weight: Bold;")
        layout.addWidget(buttonBox)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            telephon = self.Tele.text()
            for i, personne in enumerate(self.personne):
                if personne.telephon == telephon :
                    self.personne.pop(i)
                    self.table_personnes.removeRow(i)
                    break

            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            for row in range(2, sheet.max_row +1):
                cell_tel = sheet.cell(row,6)
                if int(cell_tel.value)  == int(telephon) :
                    sheet.delete_rows(row)
                    wb.save("CORONA.xlsx")
                    QtWidgets.QMessageBox.information(self, "Succès", "Succès de suppression")
                    return
        QtWidgets.QMessageBox.critical(self, "Error", "Aucune personne n'a été trouvée avec cette nationalité")

       
    def Modifier (self):
        
         # Create label and QLineEdit widget for entering CIN
        self.CIN_lab = QtWidgets.QLabel("donner le CIN de personne a modifier : ")
        self.CIN_edit = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.CIN_lab)
        layout.addWidget(self.CIN_edit)
        # Create the dialog and set its layout
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Modifier personne")
        dialog.setLayout(layout)
        # Create OK and Cancel buttons for the dialog
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        # Add the buttons to the layou
        layout.addWidget(buttonBox)
         
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # Get the CIN entered by the user
            CIN = self.CIN_edit.text()
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']

            for row in range(2, sheet.max_row +1):
                cell = sheet.cell(row,3)
                if int(cell.value) == int(CIN):
                    # Create dialog to get new information
                    layout = QtWidgets.QVBoxLayout()
                    layout.addWidget(QtWidgets.QLabel("Nouvelle adresse :"))
                    edit_adrs = QtWidgets.QLineEdit()
                    layout.addWidget(edit_adrs)
                    layout.addWidget(QtWidgets.QLabel("Nouveau numéro de téléphone :"))
                    edit_telef = QtWidgets.QLineEdit()
                    layout.addWidget(edit_telef)

                    # Create the dialog and set its layout
                    dialog = QtWidgets.QDialog(self)
                    dialog.setWindowTitle("Modifier les informations ")
                    dialog.setLayout(layout)

                    # Create OK and Cancel buttons for the dialog
                    buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
                    buttonBox.accepted.connect(dialog.accept)
                    buttonBox.rejected.connect(dialog.reject)

                    # Add the buttons to the layout
                    layout.addWidget(buttonBox)

                    # Display the dialog and wait for user input
                    if dialog.exec_() == QtWidgets.QDialog.Accepted:
                        # Update the person's information
                        adrs = edit_adrs.text()
                        telef =edit_telef.text()
                        cell=sheet.cell(row,5)
                        cell.value=adrs
                        cell=sheet.cell(row,6)
                        cell.value=telef
                        wb.save("CORONA.xlsx")
                    QtWidgets.QMessageBox.information(self, "Succès", "Les informations sont modifié")
                    return  # Exit the function once the person has been found and updated

            # If the function hasn't returned by this point, the person wasn't found
            QtWidgets.QMessageBox.critical(self, "Error", "Aucune personne n'a été trouvée avec ce CIN.")

    '''recherche et affichage'''
    def afficher_contenu_personnes(self):
        if not hasattr(self, 'table_personnes'):
            return

    # Afficher la table des personnes dans le QTableWidget "table_personnes"
        self.afficher_personnes()

        try:
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
        except FileNotFoundError:
            QtWidgets.QMessageBox.warning(self, "Attention", "Le fichier corona.xlsx n'a pas été trouvé.")
            return
        except KeyError:
            QtWidgets.QMessageBox.warning(self, "Attention", "La feuille2 n'existe pas dans le fichier corona.xlsx.")
            return
        self.table_personnes.setRowCount(df.shape[0])
        self.table_personnes.setColumnCount(df.shape[1])
        self.table_personnes.setHorizontalHeaderLabels(df.columns)
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                self.table_personnes.setItem(i, j, QtWidgets.QTableWidgetItem(str(df.iloc[i,j])))
        self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)

    def rechercher_personne_tele(self):
     # Create a dialog window to get the telephone number from the user
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Rechercher une personne")
        layout = QtWidgets.QVBoxLayout(dialog)

        telep_label = QtWidgets.QLabel("Numéro de téléphone :")
        telep_edit = QtWidgets.QLineEdit()
        telep_label.setStyleSheet("background-color: White; color: Green; font-weight: Bold;")
        layout.addWidget(telep_label)
        layout.addWidget(telep_edit)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)

        # Show the dialog and wait for the user to close it
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            telep = telep_edit.text()
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_personnes.setRowCount(df.shape[0])
            self.table_personnes.setColumnCount(df.shape[1])
            self.table_personnes.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            
            personne_trouvee = False
            i=0
            self.table_personnes.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_telep = sheet.cell(row,6)
                if str(cell_telep.value) == telep:
                    personne_trouvee = True
                    nom = sheet.cell(row,1).value
                    prenom = sheet.cell(row,2).value
                    CIN = sheet.cell(row,3).value
                    age =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    adresse = sheet.cell(row,5).value
                    ville = sheet.cell(row,6).value
                    tele = sheet.cell(row,7).value
                    date_infection = sheet.cell(row,8).value
                    Deces = str(sheet.cell(row,9).value)
                    
                    # Add the person to the table
                    self.table_personnes.setItem(i, 0, QtWidgets.QTableWidgetItem(str(nom)))
                    self.table_personnes.setItem(i, 1, QtWidgets.QTableWidgetItem(str(prenom)))
                    self.table_personnes.setItem(i, 2, QtWidgets.QTableWidgetItem(str(CIN)))
                    self.table_personnes.setItem(i, 3, QtWidgets.QTableWidgetItem(str(age)))
                    self.table_personnes.setItem(i, 4, QtWidgets.QTableWidgetItem(str(adresse)))
                    self.table_personnes.setItem(i, 5, QtWidgets.QTableWidgetItem(str(ville)))
                    self.table_personnes.setItem(i, 6, QtWidgets.QTableWidgetItem(str(tele)))
                    self.table_personnes.setItem(i, 7, QtWidgets.QTableWidgetItem(str(date_infection)))
                    self.table_personnes.setItem(i, 8, QtWidgets.QTableWidgetItem(str(Deces)))      
                    i+=1

                self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
            # If no person is found, display an error message
        if personne_trouvee == False :
            QtWidgets.QMessageBox.critical(self, "Erreur", "Aucune personne trouvée pour ce numéro de téléphone.")
            
    def Rech_IN  (self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Rechercher une personne donné")
        layout = QtWidgets.QVBoxLayout(dialog)

        cin_label = QtWidgets.QLabel("Numéro carte CIN :")
        cin_edit = QtWidgets.QLineEdit()
        cin_label.setStyleSheet("background-color: White; color: Green; font-weight: Bold;")
        layout.addWidget(cin_label)
        layout.addWidget(cin_edit)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)

        # Show the dialog and wait for the user to close it
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            CIN = cin_edit.text()
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_personnes.setRowCount(df.shape[0])
            self.table_personnes.setColumnCount(df.shape[1])
            self.table_personnes.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            
            personne_trouvee = False
            i=0
            self.table_personnes.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_cin = sheet.cell(row,3)
                if str(cell_cin.value) == CIN:
                    personne_trouvee = True
                    nom = sheet.cell(row,1).value
                    prenom = sheet.cell(row,2).value
                    CIN = sheet.cell(row,3).value
                    age =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    adresse = sheet.cell(row,5).value
                    ville = sheet.cell(row,6).value
                    tele = sheet.cell(row,7).value
                    date_infection = sheet.cell(row,8).value
                    Deces = str(sheet.cell(row,9).value)
                    
                    # Add the person to the table
                    self.table_personnes.setItem(i, 0, QtWidgets.QTableWidgetItem(str(nom)))
                    self.table_personnes.setItem(i, 1, QtWidgets.QTableWidgetItem(str(prenom)))
                    self.table_personnes.setItem(i, 2, QtWidgets.QTableWidgetItem(str(CIN)))
                    self.table_personnes.setItem(i, 3, QtWidgets.QTableWidgetItem(str(age)))
                    self.table_personnes.setItem(i, 4, QtWidgets.QTableWidgetItem(str(adresse)))
                    self.table_personnes.setItem(i, 5, QtWidgets.QTableWidgetItem(str(ville)))
                    self.table_personnes.setItem(i, 6, QtWidgets.QTableWidgetItem(str(tele)))
                    self.table_personnes.setItem(i, 7, QtWidgets.QTableWidgetItem(str(date_infection)))
                    self.table_personnes.setItem(i, 8, QtWidgets.QTableWidgetItem(str(Deces)))      
                    i+=1

                self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
            # If no person is found, display an error message
        if personne_trouvee == False :
            QtWidgets.QMessageBox.critical(self, "Erreur", "Aucune personne trouvée pour ce CIN")
        

    def rechercher_personne_nationalite(self):
    # Create a dialog window to get the nationality from the user
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Rechercher une personne")
        layout = QtWidgets.QVBoxLayout(dialog)

        nationalite_label = QtWidgets.QLabel("Nationalité :")
        nationalite_edit = QtWidgets.QLineEdit()
        nationalite_label.setStyleSheet("background-color: White; color: Green; font-weight: Bold;")
        layout.addWidget(nationalite_label)
        layout.addWidget(nationalite_edit)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        buttonBox.setStyleSheet("background-color: white; color: black; font-weight: Bold;")
        layout.addWidget(buttonBox)

    # Show the dialog and wait for the user to close it
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            nationalite = nationalite_edit.text()

    # Search for the person with the given nationality
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_personnes.setRowCount(df.shape[0])
            self.table_personnes.setColumnCount(df.shape[1])
            self.table_personnes.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            
            personne_trouvee = False
            i=0
            self.table_personnes.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_NAS = sheet.cell(row,7)
                if str(cell_NAS.value) == nationalite:
                    personne_trouvee = True
                    nom = sheet.cell(row,1).value
                    prenom = sheet.cell(row,2).value
                    CIN = sheet.cell(row,3).value
                    age =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    adresse = sheet.cell(row,5).value
                    ville = sheet.cell(row,6).value
                    tele = sheet.cell(row,7).value
                    date_infection = sheet.cell(row,8).value
                    Deces = str(sheet.cell(row,9).value)

                    # Add the person to the table
                    self.table_personnes.setItem(i, 0, QtWidgets.QTableWidgetItem(str(nom)))
                    self.table_personnes.setItem(i, 1, QtWidgets.QTableWidgetItem(str(prenom)))
                    self.table_personnes.setItem(i, 2, QtWidgets.QTableWidgetItem(str(CIN)))
                    self.table_personnes.setItem(i, 3, QtWidgets.QTableWidgetItem(str(age)))
                    self.table_personnes.setItem(i, 4, QtWidgets.QTableWidgetItem(str(adresse)))
                    self.table_personnes.setItem(i, 5, QtWidgets.QTableWidgetItem(str(ville)))
                    self.table_personnes.setItem(i, 6, QtWidgets.QTableWidgetItem(str(tele)))
                    self.table_personnes.setItem(i, 7, QtWidgets.QTableWidgetItem(str(date_infection)))
                    self.table_personnes.setItem(i, 8, QtWidgets.QTableWidgetItem(str(Deces)))      
                    i+=1
                self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
              
            # If the nationality is not found, display an error message
            QtWidgets.QMessageBox.warning(self, "Erreur", "Aucune personne trouvée de cette nationalité.")


    def rechercher_personne_decede(self):

    # Search for the person with the given nationality
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_personnes.setRowCount(df.shape[0])
            self.table_personnes.setColumnCount(df.shape[1])
            self.table_personnes.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            
            personne_trouvee = False
            i=0
            self.table_personnes.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_NAS = sheet.cell(row,9)
                if str(cell_NAS.value) == "True":
                    personne_trouvee = True
                    nom = sheet.cell(row,1).value
                    prenom = sheet.cell(row,2).value
                    CIN = sheet.cell(row,3).value
                    age =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    adresse = sheet.cell(row,5).value
                    ville = sheet.cell(row,6).value
                    tele = sheet.cell(row,7).value
                    date_infection = sheet.cell(row,8).value
                    Deces = str(sheet.cell(row,9).value)

                    # Add the person to the table
                    self.table_personnes.setItem(i, 0, QtWidgets.QTableWidgetItem(str(nom)))
                    self.table_personnes.setItem(i, 1, QtWidgets.QTableWidgetItem(str(prenom)))
                    self.table_personnes.setItem(i, 2, QtWidgets.QTableWidgetItem(str(CIN)))
                    self.table_personnes.setItem(i, 3, QtWidgets.QTableWidgetItem(str(age)))
                    self.table_personnes.setItem(i, 4, QtWidgets.QTableWidgetItem(str(adresse)))
                    self.table_personnes.setItem(i, 5, QtWidgets.QTableWidgetItem(str(ville)))
                    self.table_personnes.setItem(i, 6, QtWidgets.QTableWidgetItem(str(tele)))
                    self.table_personnes.setItem(i, 7, QtWidgets.QTableWidgetItem(str(date_infection)))
                    self.table_personnes.setItem(i, 8, QtWidgets.QTableWidgetItem(str(Deces)))      
                    i+=1
                self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        
    def rechercher_personne_non_decede(self):
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_personnes.setRowCount(df.shape[0])
            self.table_personnes.setColumnCount(df.shape[1])
            self.table_personnes.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            
            personne_trouvee = False
            i=0
            self.table_personnes.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_NAS = sheet.cell(row,9)
                if str(cell_NAS.value) == "False":
                    personne_trouvee = True
                    nom = sheet.cell(row,1).value
                    prenom = sheet.cell(row,2).value
                    CIN = sheet.cell(row,3).value
                    age =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    adresse = sheet.cell(row,5).value
                    ville = sheet.cell(row,6).value
                    tele = sheet.cell(row,7).value
                    date_infection = sheet.cell(row,8).value
                    Deces = str(sheet.cell(row,9).value)

                    # Add the person to the table
                    self.table_personnes.setItem(i, 0, QtWidgets.QTableWidgetItem(str(nom)))
                    self.table_personnes.setItem(i, 1, QtWidgets.QTableWidgetItem(str(prenom)))
                    self.table_personnes.setItem(i, 2, QtWidgets.QTableWidgetItem(str(CIN)))
                    self.table_personnes.setItem(i, 3, QtWidgets.QTableWidgetItem(str(age)))
                    self.table_personnes.setItem(i, 4, QtWidgets.QTableWidgetItem(str(adresse)))
                    self.table_personnes.setItem(i, 5, QtWidgets.QTableWidgetItem(str(ville)))
                    self.table_personnes.setItem(i, 6, QtWidgets.QTableWidgetItem(str(tele)))
                    self.table_personnes.setItem(i, 7, QtWidgets.QTableWidgetItem(str(date_infection)))
                    self.table_personnes.setItem(i, 8, QtWidgets.QTableWidgetItem(str(Deces)))      
                    i+=1

                self.table_personnes.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        


    '''gestion de maladie'''
    def mal (self):
        self.centralWidget =QtWidgets.QWidget()
        self.setCentralWidget(self.centralWidget)

        self.layout =QtWidgets.QVBoxLayout()

        self.label_code=QtWidgets.QLabel("code (numéro séquentiel) :")
        self.edit_code =QtWidgets.QLineEdit()
        self.edit_code.setMaximumWidth(350)
        
        self.label_cin = QtWidgets.QLabel("Numéro CIN : ")
        self.edit_cin = QtWidgets.QLineEdit()
        self.edit_cin.setMaximumWidth(350)

        self.label_nomm = QtWidgets.QLabel("Nom maladie :")
        self.edit_nomm = QtWidgets.QLineEdit()
        self.edit_nomm.setMaximumWidth(350)

        self.label_nb = QtWidgets.QLabel("nombre d'années :")
        self.edit_nb = QtWidgets.QLineEdit()
        self.edit_nb.setMaximumWidth(350)

        self.btn_ajouter = QtWidgets.QPushButton("Ajouter")
        self.btn_ajouter.clicked.connect(self.ajouter_maladie)
        self.btn_ajouter.setMaximumWidth(350)
        self.btn_ajouter.setStyleSheet("background-color: Crimson; color: Gold; font-weight: Bold;")

        self.table_maladies = QtWidgets.QTableWidget()
        self.table_maladies.setColumnCount(4)
        self.table_maladies.setHorizontalHeaderLabels(["Code ", "Numéro CIN :", "Nom maladie", "Nombre d'années"])
        self.layout.addWidget(self.label_code)
        self.layout.addWidget(self.edit_code)
        self.layout.addWidget(self.label_cin)
        self.layout.addWidget(self.edit_cin)
        self.layout.addWidget(self.label_nomm)
        self.layout.addWidget(self.edit_nomm)
        self.layout.addWidget(self.label_nb)
        self.layout.addWidget(self.edit_nb)
        self.centralWidget.setLayout(self.layout)
        self.layout.addWidget(self.btn_ajouter)
        self.layout.addWidget(self.table_maladies)

        self.centralWidget.setLayout(self.layout)
        self.table_maladies.setStyleSheet("background-color: Gray; color: FireBrick; font-weight: Bold;")        
        self.table_maladies.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)


    def ajouter_maladie(self):
            nomal = self.edit_nomm.text()
            CIN = self.edit_cin.text()
            code = self.edit_code.text()
            nb = self.edit_nb.text()
            
            if CIN and code  and nomal and nb :
                Maladie = maladies (CIN,code,nomal,nb)
                self.Maladie.append(Maladie)
                self.afficher_maladie()
                self.edit_nomm.setText("")
                if nomal.isalpha() ==False:
                    QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le nom de la maladie soit une chaine")
                    sys.exit() 
                self.edit_cin.setText("")
                if  CIN.isnumeric() == False:
                    QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le CIN soit un entier")
                    sys.exit()    
                
                self.edit_code.setText("")
                if  code.isnumeric() == False:
                    QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le code soit un entier")
                    sys.exit()
                
                self.edit_nb.setText("")
                if  nb.isnumeric() == False:
                    QtWidgets.QMessageBox.warning(self, "Attention", "il faut que le Nbr d'années soit un entier")
                    sys.exit()
                
                
            else:
                QtWidgets.QMessageBox.warning(self, "Attention", "Veuillez remplir tous les champs.")
                

            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F2']
            sheet["A1"] = "Nom maladie"
            sheet["B1"]= "CIN"
            sheet["C1"]="Code"
            sheet["D1"]="nombre d'années "
            
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1, value=nomal)
            sheet.cell(row=row, column=2, value=CIN)
            sheet.cell(row=row, column=3, value=code)
            sheet.cell(row=row, column=4, value=nb)

            wb.save("CORONA.xlsx")
                
    def afficher_maladie(self):
        self.table_maladies.setRowCount(len(self.Maladie))
        for i, Maladie in enumerate(self.Maladie):
            code = QtWidgets.QTableWidgetItem(Maladie.code)
            cin = QtWidgets.QTableWidgetItem(Maladie.CIN)
            nomm = QtWidgets.QTableWidgetItem(Maladie.nomal)
            nb = QtWidgets.QTableWidgetItem(Maladie.nb)
            # Set the QTableWidgetItem objects in the table
            self.table_maladies.setItem(i, 0, code)
            self.table_maladies.setItem(i, 1, cin)
            self.table_maladies.setItem(i, 2, nomm)
            self.table_maladies.setItem(i, 3, nb)
            
    def suppression_maladie(self):
        self.nomal = QtWidgets.QLabel("CIN du malade :")
        self.Nomal = QtWidgets.QLineEdit()
        self.prenomm = QtWidgets.QLabel("Nom de la maladie:")
        self.Prenomm = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.nomal)
        layout.addWidget(self.Nomal)
        layout.addWidget(self.prenomm)
        layout.addWidget(self.Prenomm)
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Supprimer personne")
        dialog.setStyleSheet("background-color: pink ; color: Maroon; font-weight: Bold;")
        dialog.setLayout(layout)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        buttonBox.setStyleSheet("background-color: purple")

        layout.addWidget(buttonBox)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            nomal = self.Nomal.text()
            prenomm = self.Prenomm.text()
            for i, Maladie in enumerate(self.Maladie):
                if Maladie.nomal == nomal and Maladie.prenomm == prenomm :
                    self.Maladie.pop(i)
                    self.table_maladies.removeRow(i)
                    break
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F2']
            for row in range(2, sheet.max_row +1):
                cell_noml = sheet.cell(row,2)
                cell_prnml= sheet.cell(row,1)
                if int(cell_noml.value)  == int(nomal) :
                    sheet.delete_rows(row)
                    wb.save("CORONA.xlsx")
                    QtWidgets.QMessageBox.information(self, "Succès", "Succès de suppression")
                    return
        QtWidgets.QMessageBox.critical(self, "Error", "Aucune personne n'a été trouvée avec cette CIN")

       

    def ANN_DC(self): 
        self.cIn_lab = QtWidgets.QLabel("donnez le CIN du personne a modifier : ")
        self.cIn_edit = QtWidgets.QLineEdit()   
        self.Ann = QtWidgets.QLabel("Modifier nombre d'anneé :")
        self.aNN = QtWidgets.QLineEdit()
        self.DC = QtWidgets.QLabel("Modifier L'etat (décé=1 /non décé=0)")
        self.dc = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.cIn_lab)
        layout.addWidget(self.cIn_edit)
        layout.addWidget(self.Ann)
        layout.addWidget(self.aNN)
        layout.addWidget(self.DC)
        layout.addWidget(self.dc)
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Modification de l'état : ")
        dialog.setStyleSheet("background-color: #E1D0C0 ;color: blue ; font-weight: Bold;")
        dialog.setLayout(layout)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            CIN = self.cIn_edit.text()
            anne=self.aNN.text()
            dece=self.dc.text()
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']
            sheet1= wb['F2']
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, 3)
                cell1= sheet1.cell(row,2)
                if  int(cell1.value) == int(CIN):
                    sheet1.cell(row, 4).value = anne

                if int(cell.value)== int(CIN):
                    if int (dece) == 1:
                        sheet.cell(row, 9).value = "True"
                    elif int(dece) == 0 :
                        sheet.cell(row, 9).value = "False"
                    else :
                        QtWidgets.QMessageBox.critical(self, "Erreur", "Il faut que le décé soit un nombre binaire 0/1 ")
                        #self.dc.setPlaceholderText('Il faut que le décé soit un nombre binaire 0/1 ')
                        

                wb.save("CORONA.xlsx")
                QtWidgets.QMessageBox.information(self, "Succès", "Les informations ont été modifiées.")
                return  # Exit the function once the person has been found and updated

        # If the function hasn't returned by this point, the person wasn't found
            QtWidgets.QMessageBox.critical(self, "Erreur", "Aucune personne n'a été trouvée avec ce CIN.")
    
    def afficher_contenu_maladie(self):
        if not hasattr(self, 'table_maladies'):
            return

    # Afficher la table des personnes dans le QTableWidget "table_personnes"
        self.afficher_maladie()

        try:
            df = pd.read_excel('CORONA.xlsx', sheet_name='F2', engine='openpyxl')
        except FileNotFoundError:
            QtWidgets.QMessageBox.warning(self, "Attention", "Le fichier corona.xlsx n'a pas été trouvé.")
            return
        except KeyError:
            QtWidgets.QMessageBox.warning(self, "Attention", "La feuille2 n'existe pas dans le fichier corona.xlsx.")
            return

        self.table_maladies.setRowCount(df.shape[0])
        self.table_maladies.setColumnCount(df.shape[1])
        self.table_maladies.setHorizontalHeaderLabels(df.columns)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                self.table_maladies.setItem(i, j, QtWidgets.QTableWidgetItem(str(df.iloc[i,j])))
        self.table_maladies.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
    
    def rechercher_maladie_personne(self):
    # Create a dialog window to get the nationality from the user
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Recherche maladie d'une personne")
        layout = QtWidgets.QVBoxLayout(dialog)

        cin_label = QtWidgets.QLabel("donner CIN :")
        cin_label.setStyleSheet("background-color: White; color: Blue; font-weight: Bold;")
        cin_edit = QtWidgets.QLineEdit()
        layout.addWidget(cin_label)
        layout.addWidget(cin_edit)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        buttonBox.setStyleSheet("background-color: Aqua; color: Blue; font-weight: Bold;")
    # Show the dialog and wait for the user to close it
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            CIN = cin_edit.text()
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_maladies.setRowCount(df.shape[0])
            self.table_maladies.setColumnCount(df.shape[1])
            self.table_maladies.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F2']
            
            personne_trouvee = False
            i=0
            self.table_maladies.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_cin = sheet.cell(row,2)
                if str(cell_cin.value) == CIN:
                    personne_trouvee = True
                    code = sheet.cell(row,1).value
                    CIN = sheet.cell(row,2).value
                    nomal = sheet.cell(row,3).value
                    nb =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    
                    # Add the person to the table
                    self.table_maladies.setItem(0, 0, QtWidgets.QTableWidgetItem(personne_trouvée.code))
                    self.table_maladies.setItem(0, 1, QtWidgets.QTableWidgetItem(personne_trouvée.CIN))
                    self.table_maladies.setItem(0, 2, QtWidgets.QTableWidgetItem(personne_trouvée.nomal))
                    self.table_maladies.setItem(0, 3, QtWidgets.QTableWidgetItem(str(personne_trouvée.nb)))
                    i+=1

                self.table_maladies.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
            # If no person is found, display an error message
        if personne_trouvee == False :
            QtWidgets.QMessageBox.critical(self, "Erreur", "Aucune personne trouvée pour ce CIN")
        

    def rechercher_par_maladie(self):
    # Create a dialog window to get the name of the disease from the user
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Recherche par une maladie")
        layout = QtWidgets.QVBoxLayout(dialog)

        nomal_label = QtWidgets.QLabel("Donner nom de la maladie:")
        nomal_label.setStyleSheet("background-color: white; color: brown; font-weight: bold;")
        nomal_edit = QtWidgets.QLineEdit()
        layout.addWidget(nomal_label)
        layout.addWidget(nomal_edit)
        nomal_label.setStyleSheet("background-color: white; color: brown; font-weight: bold;")
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        buttonBox.setStyleSheet("background-color: khaki; color: orange; font-weight: bold;")
    # Show the dialog and wait for the user to close it
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            NOM = nomal_edit.text()
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_maladies.setRowCount(df.shape[0])
            self.table_maladies.setColumnCount(df.shape[1])
            self.table_maladies.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F2']
            
            personne_trouvee = False
            i=0
            self.table_maladies.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                cell_nomal = sheet.cell(row,3)
                if str(cell_nomal.value) == NOM:
                    personne_trouvee = True
                    code = sheet.cell(row,1).value
                    CIN = sheet.cell(row,2).value
                    nomal = sheet.cell(row,3).value
                    nb =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    
                    # Add the person to the table
                    self.table_maladies.setItem(0, 0, QtWidgets.QTableWidgetItem(personne_trouvée.code))
                    self.table_maladies.setItem(0, 1, QtWidgets.QTableWidgetItem(personne_trouvée.CIN))
                    self.table_maladies.setItem(0, 2, QtWidgets.QTableWidgetItem(personne_trouvée.nomal))
                    self.table_maladies.setItem(0, 3, QtWidgets.QTableWidgetItem(str(personne_trouvée.nb)))
                    i+=1

                self.table_maladies.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
            # If no person is found, display an error message
        if personne_trouvee == False :
            QtWidgets.QMessageBox.critical(self, "Erreur", "Aucune personne trouvée pour ce CIN")
        
    def afficher_maladies_chaque_personne(self, CIN):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Recherche les maladies chaque personne")
        layout = QtWidgets.QVBoxLayout(dialog)

        cin_label = QtWidgets.QLabel("Donner cin:")
        cin_label.setStyleSheet("background-color: white; color: BlueViolet; font-weight: bold;")
        cin_edit = QtWidgets.QLineEdit()
        layout.addWidget(cin_label)
        layout.addWidget(cin_edit)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        buttonBox.setStyleSheet("background-color: khaki; color: Crimson; font-weight: bold;")

        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            CIN = cin_edit.text()
            df = pd.read_excel('CORONA.xlsx', sheet_name='F1', engine='openpyxl')
            self.table_maladies.setRowCount(df.shape[0])
            self.table_maladies.setColumnCount(df.shape[1])
            self.table_maladies.setHorizontalHeaderLabels(df.columns)
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F2']
            
            personne_trouvee = False
            i=0
            self.table_maladies.clearContents()
        # Search for people with the given phone number
            for row in range(2, sheet.max_row +1):
                print(sheet.max_row)
                cell_cin = sheet.cell(row,2)
                if str(cell_cin.value) == CIN:
                    personne_trouvee = True
                    code = sheet.cell(row,1).value
                    CIN = sheet.cell(row,2).value
                    nomal = sheet.cell(row,3).value
                    nb =str(sheet.cell(row,4).value) # Convert age to a string before setting it as QTableWidgetItem
                    
                    # Add the person to the table
                    self.table_maladies.setItem(0, 0, QtWidgets.QTableWidgetItem(personne_trouvée.code))
                    self.table_maladies.setItem(0, 1, QtWidgets.QTableWidgetItem(personne_trouvée.CIN))
                    self.table_maladies.setItem(0, 2, QtWidgets.QTableWidgetItem(personne_trouvée.nomal))
                    self.table_maladies.setItem(0, 3, QtWidgets.QTableWidgetItem(str(personne_trouvée.nb)))
                    i+=1

                self.table_maladies.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
            # If no person is found, display an error message
        if personne_trouvee == False :
            QtWidgets.QMessageBox.critical(self, "Erreur", "Aucune personne trouvée pour ce CIN")
        

    def rechercher_pourcentage_maladies(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Recherche pourcentage d'une maladie")
        layout = QtWidgets.QVBoxLayout(dialog)

        nomal_label = QtWidgets.QLabel("Donner nom de la maladie:")
        nomal_label.setStyleSheet("background-color: white; color: Red; font-weight: bold;")
        nomal_edit = QtWidgets.QLineEdit()
        layout.addWidget(nomal_label)
        layout.addWidget(nomal_edit)
        nomal_edit.setStyleSheet("background-color: white; color: Pink; font-weight: bold;")
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        buttonBox.setStyleSheet("background-color: khaki; color: RedOrange; font-weight: bold;")
        num_people = 0
        # Get the disease name entered by the user
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            NOM = nomal_edit.text()
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F2']
            
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, 1)
                if str(cell.value) == NOM:
                     num_people+= 1

            # Calculate the percentage of people for the given disease
            total_people = sheet.max_row - 1
            percentage = (num_people / total_people) * 100

            # Display the results in a message box
            message = f"Pourcentage de personnes atteintes de {NOM} : {percentage:.2f}%"

            #QtWidgets.QMessageBox.information(self, f"Pourcentage de {nomal}", message)
            messageBox = QtWidgets.QMessageBox(self)
            messageBox.setWindowTitle(f"Pourcentage de {NOM}")
            messageBox.setText(f"Pourcentage de personnes atteintes de {NOM} : {percentage:.2f}%")
            messageBox.setStyleSheet("QMessageBox { background-color: Chartreuse;font-weight: bold; }")
            messageBox.exec_()


    '''calcul et affichage'''
    def afficher_personnes_par_ville(self):
        ville_label = QtWidgets.QLabel("Ville :")
        ville_edit = QtWidgets.QLineEdit()
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(ville_label)
        layout.addWidget(ville_edit)
        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        layout.addWidget(buttonBox)
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Afficher les personnes de cette ville:")
        dialog.setLayout(layout)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)

        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # Ouvrir le classeur Excel et accéder à la feuille table_personnes
            wb = openpyxl.load_workbook("CORONA.xlsx")
            sheet = wb['F1']

            # Afficher les en-têtes de colonnes
            headers = ['Nom', 'Prénom', 'CIN', 'Age', 'Adresse', 'Téléphone', 'Nationalité', "Date d'infection", 'Décédé']
            self.table_personnes.setColumnCount(len(headers))
            self.table_personnes.setHorizontalHeaderLabels(headers)

            # Parcourir les lignes de la feuille et ajouter ceux ayant la ville donnée
            row_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                nom, prenom, CIN, age, adresse, ville, tele, date_infection, decede = row
                if ville == ville_edit.text():
                    self.table_personnes.insertRow(row_count)
                    self.table_personnes.setItem(row_count, 0, QtWidgets.QTableWidgetItem(nom))
                    self.table_personnes.setItem(row_count, 1, QtWidgets.QTableWidgetItem(prenom))
                    self.table_personnes.setItem(row_count, 2, QtWidgets.QTableWidgetItem(str(CIN)))
                    self.table_personnes.setItem(row_count, 3, QtWidgets.QTableWidgetItem(str(age)))
                    self.table_personnes.setItem(row_count, 4, QtWidgets.QTableWidgetItem(adresse))
                    self.table_personnes.setItem(row_count, 5, QtWidgets.QTableWidgetItem(str(tele)))
                    self.table_personnes.setItem(row_count, 6, QtWidgets.QTableWidgetItem(ville))
                    self.table_personnes.setItem(row_count, 7, QtWidgets.QTableWidgetItem(date_infection.strftime('%d/%m/%Y')))
                    self.table_personnes.setItem(row_count, 8, QtWidgets.QTableWidgetItem(str(decede)))
                    row_count += 1

            if row_count == 0:
                # Si aucune personne n'est trouvée, afficher un message d'erreur
                QtWidgets.QMessageBox.warning(self, "Attention", "Aucune personne trouvée pour cette ville.")


    '''rcupération et enregistrement'''
    def enregistrer_personne_excel(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Enregistrer les données des personnes")
        layout = QtWidgets.QVBoxLayout(dialog)

        enr_label = QtWidgets.QLabel("Enregistrer les données ?")
        enr_label.setStyleSheet("background-color: pink; color: red; font-weight: bold;")
        layout.addWidget(enr_label)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        buttonBox.setStyleSheet("background-color: khaki; color: RedOrange; font-weight: bold;")

        result = dialog.exec_()
        if result == QtWidgets.QDialog.Accepted:
            # Charger le classeur Excel
            wb = openpyxl.load_workbook("CORONA.xlsx")

            # Accéder à la feuille F1
            sheet = wb['F1']

            # Enregistrer les données des personnes
            for i, personne in enumerate(self.personne):
                sheet.cell(row=i+2, column=1, value=personne.nom)
                sheet.cell(row=i+2, column=2, value=personne.prenom)
                sheet.cell(row=i+2, column=3, value=personne.CIN)
                sheet.cell(row=i+2, column=4, value=personne.age)
                sheet.cell(row=i+2, column=5, value=personne.adrs)
                sheet.cell(row=i+2, column=6, value=personne.villle)
                sheet.cell(row=i+2, column=7, value=personne.tele)
                sheet.cell(row=i+2, column=8, value=personne.date_infection)
                sheet.cell(row=i+2, column=9, value=personne.décé)

            # Enregistrer les modifications dans le fichier Excel
            wb.save("CORONA.xlsx")

            # Afficher un message de succès
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Information)
            msgBox.setWindowTitle("Succès")
            msgBox.setText("Les données ont été enregistrées avec succès.")
            msgBox.setStyleSheet("background-color: Pink; color: RedOrange; font-weight: bold;")
            msgBox.exec_()

    

    def enregistrer_maladie_excel(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Enregistrer les données MALADIES")
        layout = QtWidgets.QVBoxLayout(dialog)

        enr_label = QtWidgets.QLabel("Enregistrer les données MALADIES ?")
        enr_label.setStyleSheet("background-color: Yellow; color: red; font-weight: bold;")
        layout.addWidget(enr_label)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)
        buttonBox.setStyleSheet("background-color: Pink; color: RedOrange; font-weight: bold;")

        result = dialog.exec_()
        if result == QtWidgets.QDialog.Accepted:
            # Charger le classeur Excel
            wb = openpyxl.load_workbook("CORONA.xlsx")

            # Accéder à la feuille F1
            sheet = wb['F2']

            # Enregistrer les données des personnes
            for i, Maladies in enumerate(self.Maladie):
                sheet.cell(row=i+2, column=1, value=Maladies.nom)
                sheet.cell(row=i+2, column=2, value=Maladies.CIN)
                sheet.cell(row=i+2, column=3, value=Maladies.nomal)
                sheet.cell(row=i+2, column=4, value=Maladies.nb)
                

            # Enregistrer les modifications dans le fichier Excel
            wb.save("CORONA.xlsx")

            # Afficher un message de succès
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Information)
            msgBox.setWindowTitle("Succès")
            msgBox.setText("Les données ont été enregistrées avec succès.")
            msgBox.setStyleSheet("background-color: Pink; color: RedOrange; font-weight: bold;")
            msgBox.exec_()



    def recuperer_personne_excel(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Récuperer les données des Personnes")
        layout = QtWidgets.QVBoxLayout(dialog)

        r_label = QtWidgets.QLabel("Récuperer les données PERSONNES ?")
        r_label.setStyleSheet("background-color: Yellow; color: red; font-weight: bold;")
        layout.addWidget(r_label)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        buttonBox.setStyleSheet("background-color: Pink; color: RedOrange; font-weight: bold;")
        layout.addWidget(buttonBox)
        

        # Charger le classeur Excel
        wb = openpyxl.load_workbook("CORONA.xlsx")

        # Accéder à la feuille F1
        sheet = wb['F1']

        # Récupérer les données des personnes
        personnes = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nom, prenom, CIN, age, adresse, ville, tele, date_infection, décé = row
            personne = Personne(nom, prenom, CIN, age, adresse, ville, tele, date_infection, décé)
            personnes.append(personne)

        # Afficher le message de succès
        msgBox = QtWidgets.QMessageBox()
        msgBox.setIcon(QtWidgets.QMessageBox.Information)
        msgBox.setWindowTitle("Succès")
        msgBox.setText("Les données ont été récupérées avec succès.")
        msgBox.setStyleSheet("background-color: DeepPink; color: RedOrange; font-weight: bold;")
        msgBox.exec_()

        return personnes


    def recuperer_maladie_excel(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Récupérer les données des Maladies")
        layout = QtWidgets.QVBoxLayout(dialog)

        r_label = QtWidgets.QLabel("Récupérer les données des Maladies ?")
        r_label.setStyleSheet("background-color: Yellow; color: red; font-weight: bold;")
        layout.addWidget(r_label)

        buttonBox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        buttonBox.setStyleSheet("color: RedOrange; font-weight: bold; ")
        layout.addWidget(buttonBox)

        # Charger le classeur Excel
        wb = openpyxl.load_workbook("CORONA.xlsx")

        # Accéder à la feuille F2
        sheet = wb['F2']

        # Récupérer les données des maladies
        maladi = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code,CIN, nomal, nb = row
            maladie = maladies(code,CIN, nomal, nb)
            maladi.append(maladie)

        # Afficher le message de succès
        msgBox = QtWidgets.QMessageBox()
        msgBox.setIcon(QtWidgets.QMessageBox.Information)
        msgBox.setWindowTitle("Succès")
        msgBox.setText("Les données ont été récupérées avec succès.")
        msgBox.setStyleSheet("background-color: Violet; color: RedOrange; font-weight: bold;")
        msgBox.exec_()

        return maladies

    
def window():
    app=QtWidgets.QApplication(sys.argv)
    win= my_window()#ou bien without class :QtWidgets.QMainWindow()creation of desktop window
    win.show()
    sys.exit(app.exec_())
    
window() #appel de la fonction